#!/usr/bin/env python3
"""
从 i18n/strings.xlsx 生成各模块的 values/strings.xml 和 values-{locale}/strings.xml。

Excel 格式：
  - 一个 sheet 对应一个 Android 模块
  - 首行：语言代码（首列固定为 "key"）
  - 后续行：key 名及各语言翻译

使用方法：
  pip install openpyxl
  python scripts/generate-i18n-strings.py
"""

import os
import xml.sax.saxutils as saxutils

from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = PROJECT_ROOT / "i18n" / "strings.xlsx"

# Sheet 名称 → 模块 res 目录的相对路径
SHEET_TO_RES = {
    "app": "app/src/main/res",
    "feature-home-impl": "feature/home/impl/src/main/res",
    "feature-settings-impl": "feature/settings/impl/src/main/res",
}


def escape_xml(text: str) -> str:
    """转义 XML 特殊字符，保留已有的 XML 实体。"""
    return saxutils.escape(text)


def generate_strings_xml(keys: list[str], values: list[str]) -> str:
    """生成 strings.xml 内容。"""
    lines = ['<?xml version="1.0" encoding="utf-8"?>', "<resources>"]

    # 按 key 排序，生成稳定的输出
    items = sorted(zip(keys, values), key=lambda x: x[0])
    for key, value in items:
        escaped = escape_xml(value)
        if "%" in escaped:
            lines.append(f'    <string name="{key}">{escaped}</string>')
        else:
            lines.append(f'    <string name="{key}">{escaped}</string>')
    lines.append("</resources>")
    return "\n".join(lines) + "\n"


def process_sheet(sheet):
    """处理一个 sheet，返回 {locale: [(key, value), ...]}。"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    # 首行：语言代码
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not headers or headers[0].lower() != "key":
        print(f"  ⚠ 跳过：首列不是 'key'，实际为 '{headers[0] if headers else ''}'")
        return {}

    locales = headers[1:]  # ["zh", "en", ...]
    result: dict[str, list[tuple[str, str]]] = {loc: [] for loc in locales}

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        if not key:
            continue
        for i, locale in enumerate(locales):
            if i + 1 < len(row) and row[i + 1] is not None:
                value = str(row[i + 1]).strip()
                if value:
                    result[locale].append((key, value))

    return result


def write_res_files(res_dir: Path, locale_data: dict[str, list[tuple[str, str]]]):
    """
    写入 strings.xml：
      - 第一个 locale 写入 values/strings.xml（默认）
      - 其他 locale 写入 values-{locale}/strings.xml
    """
    locales = list(locale_data.keys())
    if not locales:
        return

    # 第一个 locale 作为默认语言
    default_locale = locales[0]
    default_path = res_dir / "values" / "strings.xml"
    default_path.parent.mkdir(parents=True, exist_ok=True)
    keys = [k for k, v in locale_data[default_locale]]
    vals = [v for k, v in locale_data[default_locale]]
    content = generate_strings_xml(keys, vals)
    default_path.write_text(content, encoding="utf-8")
    print(f"  ✅ 已生成: {default_path.relative_to(PROJECT_ROOT)} ({len(keys)} strings)")

    # 其他 locale
    for locale in locales[1:]:
        locale_dir = res_dir / f"values-{locale}"
        locale_dir.mkdir(parents=True, exist_ok=True)
        locale_path = locale_dir / "strings.xml"
        keys = [k for k, v in locale_data[locale]]
        vals = [v for k, v in locale_data[locale]]
        content = generate_strings_xml(keys, vals)
        locale_path.write_text(content, encoding="utf-8")
        print(f"  ✅ 已生成: {locale_path.relative_to(PROJECT_ROOT)} ({len(keys)} strings)")


def main():
    if not XLSX_PATH.exists():
        print(f"❌ 未找到 Excel 文件: {XLSX_PATH}")
        print("请先创建 i18n/strings.xlsx，每个模块一个 sheet。")
        return

    try:
        import openpyxl
    except ImportError:
        print("❌ 需要 openpyxl 库，请运行: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_RES:
            print(f"  ⚠ 跳过未知 sheet: '{sheet_name}'（未在映射表中定义）")
            continue

        res_path = PROJECT_ROOT / SHEET_TO_RES[sheet_name]
        print(f"\n📄 Sheet: {sheet_name} → {SHEET_TO_RES[sheet_name]}")

        sheet = wb[sheet_name]
        locale_data = process_sheet(sheet)
        if not locale_data:
            print(f"  ⚠ 无有效数据，跳过")
            continue

        write_res_files(res_path, locale_data)

    print("\n✅ 全部生成完成！")


if __name__ == "__main__":
    main()
